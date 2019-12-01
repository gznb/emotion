from django.http import JsonResponse, HttpResponseServerError, HttpResponse
from d2Result.utils.check_get_data import CheckReceiveFormat
from rest_framework.views import APIView
from openpyxl import Workbook
import os
from pymongo_aggregate.aggregate_rule import AggregateRule
from d2Result.models import d2ResultModel
from conf.field_conf import SEARCH_DICT

import logging
logger = logging.getLogger('django')


class ExportView(APIView):

    def __init__(self):
        self.aggregate_rules = []
        super().__init__()
    def check_format(self, request):
        order_id = request.data.get('orderId')
        data = request.data.get('data')
        this_period = data.get('thisPeriod')
        source = data.get('source')
        word = data.get('word')
        search = data.get('search')
        attribute = data.get('attribute')
        check = CheckReceiveFormat()
        try:
            order_id = check.check_order_id({'orderId': order_id})

            begin_time, end_time = check.check_period(this_period)

            s_is_all, s_count, s_list = check.check_source(source)
            w_is_all, w_count, w_list = check.check_word(word)
            is_default, key, value = check.check_search(search)
            attribute_is_default, attribute_value = check.check_attribute_search(attribute)
        except (TypeError, ValueError) as err:
            return JsonResponse({'code': 2, 'msg': str(err), 'data': {}})
        except Exception as err:
            logger.error(err)
            return HttpResponseServerError()
        else:
            rules = AggregateRule()
            self.aggregate_rules.append(rules.equal_rule('GorderId', order_id))
            # 将属性判断加在前面
            if attribute_is_default != 1:
                self.aggregate_rules.append(rules.equal_rule("GresultAttribute", attribute_value))
            self.aggregate_rules.append(rules.time_interval_rule('GresultReleaseTime', begin_time, end_time))
            if w_is_all != 1:
                self.aggregate_rules.append(rules.list_in_rule('GresultKeyword', w_list))
            self.aggregate_rules.append(rules.look_up_role('d2_spider_model','GspiderId', 'GspiderId', 'spider'))
            if s_is_all != 1:
                self.aggregate_rules.append(rules.list_in_rule('spider.GspiderClassification', s_list))
            # print(is_default)
            if is_default != 1:
                if '渠道' in key:
                    key = 'spider.' + SEARCH_DICT[key]
                else:
                    # print(key)
                    key = SEARCH_DICT[key]
                self.aggregate_rules.append(rules.re_rule(key, value))
            # 最多下载1000条
            self.aggregate_rules = self.aggregate_rules + rules.skip_limit(1, 1000)

    def post(self, request, *args, **kwargs):
        # print(1111)
        get_data = self.check_format(request)
        # print(2222)
        if isinstance(get_data, (JsonResponse, HttpResponseServerError)):
            return get_data
        else:

            # print(self.aggregate_rules)
            rev_data = d2ResultModel._get_collection().aggregate(self.aggregate_rules)
            ws = Workbook()
            w = ws.create_sheet(u"舆情分析表", 0)
            w.cell(row=1, column=1).value = "url"
            w.cell(row=1, column=2).value = "标题"
            w.cell(row=1, column=3).value = "属性"
            w.cell(row=1, column=4).value = "渠道"
            w.cell(row=1, column=5).value = "正文"
            w.cell(row=1, column=6).value = "爬取时间"
            w.cell(row=1, column=7).value = "发布时间"
            w.cell(row=1, column=8).value = "检测词"
            file_name = 'data_one.xlsx'
            exist_file = os.path.exists(file_name)
            if exist_file:
                os.remove(file_name)
            excel_row = 2
            for rev in rev_data:
                w.cell(row=excel_row, column=1).value = rev['GresultRealUrl']
                w.cell(row=excel_row, column=2).value = rev['GresultTitle'].strip()
                w.cell(row=excel_row, column=3).value = rev['GresultAttribute']
                w.cell(row=excel_row, column=4).value = rev['spider'][0]['GspiderClassification']
                w.cell(row=excel_row, column=5).value = rev['GresultDetailedInformating'].strip() if rev['GresultDetailedInformating'] else "此渠道不包括正文"
                w.cell(row=excel_row, column=6).value = rev['GresultNowTime']
                w.cell(row=excel_row, column=7).value = rev['GresultReleaseTime']
                w.cell(row=excel_row, column=8).value = rev['GresultKeyword']
                excel_row += 1
            response = HttpResponse(content_type='application/msexcel')
            response['Content-Disposition'] = 'attachment;filename={}'.format(file_name)
            ws.save(response)
            return response
